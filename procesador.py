"""
Procesador de Reposición de Inventario
Genera un .xlsx con 6 hojas siguiendo exactamente el formato de referencia.
"""
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta exacta del archivo de referencia ───────────────────────────────────
C_HDR_DARK  = "1F4E79"   # azul oscuro — encabezados
C_HDR_LIGHT = "BDD7EE"   # azul claro  — total saldo / total reponer
C_ROW_A     = "EBF3FA"   # fila par
C_ROW_B     = "FFFFFF"   # fila impar
C_REPONER   = "D6E4F0"   # fila Reponer en Rep+Saldo
C_SALDO_ROW = "DEE4F1"   # fila Saldo   en Rep+Saldo
C_YELLOW    = "FFFF00"   # selector días
C_MAPA2_HD  = "FFF2CC"   # cabecera y celdas Mapa2
C_MAPA2_DIN = "FFE699"   # columna dinámica sugerido Mapa2
C_OK        = "E2EFDA"
C_BAJO      = "FCE4D6"
C_SIN_VENTA = "D9D9D9"

def _fill(c):      return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Arial")
def _center():     return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():       return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def _border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, coord, val, size=10):
    ws[coord].value = val
    ws[coord].font  = _font(bold=True, color="FFFFFF", size=size)
    ws[coord].fill  = _fill(C_HDR_DARK)
    ws[coord].alignment = _center()
    ws[coord].border    = _border()

def _row_fill(i):  return C_ROW_A if i % 2 == 0 else C_ROW_B

def _estado_color(e):
    return {"OK": C_OK, "BAJO": C_BAJO, "BAJA ROTACIÓN": C_BAJO,
            "SUGERIDO": C_MAPA2_HD, "SIN VENTA": C_SIN_VENTA}.get(e, C_SIN_VENTA)

# ── Parser ────────────────────────────────────────────────────────────────────
def parse_xls(filepath):
    engine = "openpyxl" if filepath.endswith(".xlsx") else "xlrd"
    df = pd.read_excel(filepath, engine=engine, header=None)
    branch_starts = {}
    for col_idx, val in enumerate(df.iloc[1, :]):
        if pd.notna(val) and str(val).strip() and col_idx > 1:
            branch_starts[str(val).strip()] = col_idx
    products = []
    for row_idx in range(3, len(df)):
        item = df.iloc[row_idx, 0]
        desc = df.iloc[row_idx, 1]
        if pd.isna(item) or str(item).startswith('R(s)'):
            continue
        prod = {'item': str(item).strip(), 'desc': str(desc).strip() if pd.notna(desc) else ''}
        for bname, bcol in branch_starts.items():
            pvtas = df.iloc[row_idx, bcol + 1]
            saldo = df.iloc[row_idx, bcol + 4]
            prod[f'{bname}__pvtas'] = float(pvtas) if pd.notna(pvtas) else 0.0
            prod[f'{bname}__saldo'] = float(saldo) if pd.notna(saldo) else 0.0
        products.append(prod)
    return products, list(branch_starts.keys())

# ── Helpers de cálculo ────────────────────────────────────────────────────────
def _avg_pvtas(p, branches):
    vals = [p[f'{b}__pvtas'] for b in branches if p[f'{b}__pvtas'] > 0]
    return round(np.mean(vals), 2) if vals else 0.0

def _locales_venta(p, branches):
    return sum(1 for b in branches if p[f'{b}__pvtas'] > 0)

def _estado_prod(pvtas_total, saldo_total, dias):
    if pvtas_total == 0: return "SIN VENTA"
    cob = saldo_total / (pvtas_total / 30)
    if cob >= dias: return "OK"
    if cob >= dias * 0.5: return "BAJA ROTACIÓN"
    return "BAJO"

def _estado_suc(pvt, sal, dias):
    if pvt == 0: return "SIN VENTA"
    cob = sal / (pvt / 30)
    if cob >= dias: return "OK"
    if cob >= dias * 0.5: return "SUGERIDO"
    return "BAJO"

# ── HOJA 1: Consolidado Reposición ───────────────────────────────────────────
def sheet_consolidado(wb, products, branches, dias):
    ws = wb.create_sheet("Consolidado Reposición")
    N = len(branches)

    ws["A1"].value = "Días de cálculo:"; ws["A1"].font = _font(bold=True, size=11)
    ws["B1"].value = dias
    ws["B1"].font  = _font(bold=True, color="0000FF", size=11)
    ws["B1"].fill  = _fill(C_YELLOW); ws["B1"].alignment = _center()
    ws["C1"].value = "← Cambia este valor para recalcular (30, 45 o 60 días)"
    ws["C1"].font  = _font(color="7F7F7F", size=10)

    _hdr(ws, "A3", "Código"); _hdr(ws, "B3", "Descripción")
    ws.row_dimensions[3].height = 34.5

    for bi, b in enumerate(branches):
        _hdr(ws, f"{get_column_letter(3+bi)}3", b)

    mapa2_col  = 3 + N
    total_col  = mapa2_col + 1
    vprom_col  = total_col + 1
    estado_col = vprom_col + 1

    ws[f"{get_column_letter(mapa2_col)}3"].value = "11 Mapa 2 (Sugerido)"
    ws[f"{get_column_letter(mapa2_col)}3"].font  = _font(bold=True, color="7F4F00", size=10)
    ws[f"{get_column_letter(mapa2_col)}3"].fill  = _fill(C_MAPA2_HD)
    ws[f"{get_column_letter(mapa2_col)}3"].alignment = _center()
    ws[f"{get_column_letter(mapa2_col)}3"].border    = _border()

    _hdr(ws, f"{get_column_letter(total_col)}3",  "Total Reponer")
    _hdr(ws, f"{get_column_letter(vprom_col)}3",  "V.Prom/Mes")
    _hdr(ws, f"{get_column_letter(estado_col)}3", "Estado")

    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 48
    for ci in range(3, estado_col+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    for pi, p in enumerate(products):
        row = pi + 4; rf = _row_fill(pi)
        pvt_tot = sum(p[f'{b}__pvtas'] for b in branches)
        sal_tot = sum(p[f'{b}__saldo'] for b in branches)

        def dc(col, val, bold=False, fill=None, align="center", left=False):
            c = ws.cell(row, col, val)
            c.font = _font(bold=bold, size=11)
            c.fill = _fill(fill or rf)
            c.alignment = _left() if left else _center()
            c.border = _border()

        dc(1, p['item'], bold=True)
        dc(2, p['desc'], left=True)

        for bi, b in enumerate(branches):
            pvt = p[f'{b}__pvtas']; sal = p[f'{b}__saldo']
            val = f"=MAX(0,ROUNDUP({pvt}*$B$1/30,0)-{int(sal)})" if pvt > 0 else 0
            dc(3+bi, val)

        avg = _avg_pvtas(p, branches)
        mc_val = f"=MAX(0,ROUNDUP({avg}*$B$1/30,0)-0)" if avg > 0 else "-"
        c = ws.cell(row, mapa2_col, mc_val)
        c.font = _font(bold=True, color="7F4F00", size=11)
        c.fill = _fill(C_MAPA2_HD); c.alignment = _center(); c.border = _border()

        sc = get_column_letter(3); ec = get_column_letter(2+N); mcl = get_column_letter(mapa2_col)
        tc = ws.cell(row, total_col, f"=SUM({sc}{row}:{ec}{row})+IF(ISNUMBER({mcl}{row}),{mcl}{row},0)")
        tc.font = _font(bold=True, size=11); tc.fill = _fill(C_HDR_LIGHT)
        tc.alignment = _center(); tc.border = _border()

        dc(vprom_col, int(pvt_tot) if pvt_tot == int(pvt_tot) else pvt_tot)

        est = _estado_prod(pvt_tot, sal_tot, dias)
        ec2 = ws.cell(row, estado_col, est)
        ec2.font = _font(bold=True, size=11); ec2.fill = _fill(_estado_color(est))
        ec2.alignment = _center(); ec2.border = _border()

    ws.freeze_panes = "A4"

# ── HOJA 2: Rep + Saldo + Cob + PVtas ────────────────────────────────────────
def sheet_rep_saldo(wb, products, branches, dias):
    ws = wb.create_sheet("Rep + Saldo + Cob + PVtas")
    N = len(branches)

    ws["A1"].value = "Días de cálculo:"; ws["A1"].font = _font(bold=True, size=11)
    ws["B1"].value = "='Consolidado Reposición'!B1"
    ws["B1"].font  = _font(bold=True, color="0000FF", size=11)
    ws["B1"].fill  = _fill(C_YELLOW); ws["B1"].alignment = _center()

    _hdr(ws, "A3", "Código"); _hdr(ws, "B3", "Descripción"); _hdr(ws, "C3", "Fila")
    ws.row_dimensions[3].height = 34.5

    for bi, b in enumerate(branches):
        _hdr(ws, f"{get_column_letter(4+bi)}3", b)

    mapa2_col  = 4 + N
    total_col  = mapa2_col + 1
    estado_col = total_col + 1

    ws[f"{get_column_letter(mapa2_col)}3"].value = "11 Mapa 2"
    ws[f"{get_column_letter(mapa2_col)}3"].font  = _font(bold=True, color="7F4F00", size=10)
    ws[f"{get_column_letter(mapa2_col)}3"].fill  = _fill(C_MAPA2_HD)
    ws[f"{get_column_letter(mapa2_col)}3"].alignment = _center()
    ws[f"{get_column_letter(mapa2_col)}3"].border    = _border()
    _hdr(ws, f"{get_column_letter(total_col)}3",  "Total")
    _hdr(ws, f"{get_column_letter(estado_col)}3", "Estado")

    ws.column_dimensions["A"].width = 11; ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    for ci in range(4, estado_col+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    REF = "'Consolidado Reposición'!$B$1"
    s_col = get_column_letter(4); e_col = get_column_letter(3+N); mcl = get_column_letter(mapa2_col)

    for pi, p in enumerate(products):
        base = pi * 3 + 4; rf = _row_fill(pi)
        pvt_tot = sum(p[f'{b}__pvtas'] for b in branches)
        sal_tot = sum(p[f'{b}__saldo'] for b in branches)
        avg = _avg_pvtas(p, branches)
        est = _estado_prod(pvt_tot, sal_tot, dias)

        def sc(row, col, val, fill, bold=False, left=False, color="000000"):
            c = ws.cell(row, col, val)
            c.font = _font(bold=bold, color=color, size=11)
            c.fill = _fill(fill); c.alignment = _left() if left else _center()
            c.border = _border()

        r = base
        sc(r,1,p['item'],C_ROW_A,bold=True); sc(r,2,p['desc'],C_ROW_A,left=True,bold=True)
        sc(r,3,"📦 Reponer",C_REPONER,bold=True,left=True)
        for bi,b in enumerate(branches):
            pvt=p[f'{b}__pvtas']; sal=p[f'{b}__saldo']
            val=f"=MAX(0,ROUNDUP({pvt}*{REF}/30,0)-{int(sal)})" if pvt>0 else "-"
            sc(r,4+bi,val,C_REPONER)
        m_val=f"=MAX(0,ROUNDUP({avg}*{REF}/30,0)-0)" if avg>0 else "-"
        mc=ws.cell(r,mapa2_col,m_val); mc.font=_font(bold=True,color="7F4F00",size=11)
        mc.fill=_fill(C_MAPA2_HD); mc.alignment=_center(); mc.border=_border()
        tc=ws.cell(r,total_col,f"=SUM({s_col}{r}:{e_col}{r})")
        tc.font=_font(bold=True,size=11); tc.fill=_fill(C_HDR_LIGHT); tc.alignment=_center(); tc.border=_border()
        ec=ws.cell(r,estado_col,est); ec.font=_font(bold=True,size=11)
        ec.fill=_fill(_estado_color(est)); ec.alignment=_center(); ec.border=_border()

        r = base+1
        ws.cell(r,1).border=_border(); ws.cell(r,2).border=_border()
        sc(r,3,"📊 Saldo bodega",C_SALDO_ROW,bold=True,left=True)
        for bi,b in enumerate(branches):
            sal=p[f'{b}__saldo']
            sc(r,4+bi,int(sal) if sal else "-",C_SALDO_ROW)
        sc(r,mapa2_col,"-",C_MAPA2_HD)
        tc2=ws.cell(r,total_col,f"=SUM({s_col}{r}:{e_col}{r})")
        tc2.font=_font(bold=True,size=11); tc2.fill=_fill(C_HDR_LIGHT); tc2.alignment=_center(); tc2.border=_border()
        ws.cell(r,estado_col).border=_border()

        r = base+2
        ws.cell(r,1).border=_border(); ws.cell(r,2).border=_border()
        sc(r,3,"📈 P.Vtas/Mes",rf,bold=True,left=True)
        pvt_shown=0
        for bi,b in enumerate(branches):
            pvt=p[f'{b}__pvtas']; pvt_shown+=pvt
            sc(r,4+bi,int(pvt) if pvt>0 else "-",rf)
        sc(r,mapa2_col,round(avg,1) if avg>0 else "-",C_MAPA2_HD)
        tv=int(pvt_shown) if pvt_shown==int(pvt_shown) else pvt_shown
        tc3=ws.cell(r,total_col,tv); tc3.font=_font(bold=True,size=11)
        tc3.fill=_fill(C_HDR_LIGHT); tc3.alignment=_center(); tc3.border=_border()
        ws.cell(r,estado_col).border=_border()

    ws.freeze_panes = "D4"

# ── HOJA 3: Detalle Expandido ─────────────────────────────────────────────────
def sheet_detalle_expandido(wb, products, branches, dias):
    ws = wb.create_sheet("Detalle Expandido")
    N = len(branches)

    ws["A1"].value = "Días de cálculo:"; ws["A1"].font = _font(bold=True, size=11)
    ws["B1"].value = "='Consolidado Reposición'!B1"
    ws["B1"].font  = _font(bold=True, color="0000FF", size=11)
    ws["B1"].fill  = _fill(C_YELLOW); ws["B1"].alignment = _center()

    _hdr(ws, "A2", "Código"); _hdr(ws, "B2", "Descripción")
    ws.row_dimensions[2].height = 30

    for bi, b in enumerate(branches):
        sc = 3 + bi*3; ec = sc+2
        ws.merge_cells(f"{get_column_letter(sc)}2:{get_column_letter(ec)}2")
        c = ws[f"{get_column_letter(sc)}2"]
        c.value=b; c.font=_font(bold=True,color="FFFFFF",size=10)
        c.fill=_fill(C_HDR_DARK); c.alignment=_center(); c.border=_border()

    mapa2_start = 3 + N*3
    ws.merge_cells(f"{get_column_letter(mapa2_start)}2:{get_column_letter(mapa2_start+2)}2")
    mc2 = ws[f"{get_column_letter(mapa2_start)}2"]
    mc2.value="11 Mapa 2"; mc2.font=_font(bold=True,color="7F4F00",size=10)
    mc2.fill=_fill(C_MAPA2_HD); mc2.alignment=_center(); mc2.border=_border()

    total_col = mapa2_start+3; estado_col = total_col+1
    _hdr(ws,f"{get_column_letter(total_col)}2","Total Reponer")
    _hdr(ws,f"{get_column_letter(estado_col)}2","Estado")

    _hdr(ws,"A3","Código"); _hdr(ws,"B3","Descripción")
    for bi in range(N):
        sc = 3+bi*3
        for j,sub in enumerate(["Reponer","Saldo","P.Vtas/Mes"]):
            c=ws[f"{get_column_letter(sc+j)}3"]
            c.value=sub; c.font=_font(bold=True,color="FFFFFF",size=10)
            c.fill=_fill(C_HDR_DARK); c.alignment=_center(); c.border=_border()
    for j,sub in enumerate(["Reponer","Saldo","P.Vtas/Mes"]):
        c=ws[f"{get_column_letter(mapa2_start+j)}3"]
        c.value=sub; c.font=_font(bold=True,color="7F4F00",size=10)
        c.fill=_fill(C_MAPA2_HD); c.alignment=_center(); c.border=_border()
    _hdr(ws,f"{get_column_letter(total_col)}3","Total Reponer")
    _hdr(ws,f"{get_column_letter(estado_col)}3","Estado")

    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=48
    for ci in range(3,estado_col+1):
        ws.column_dimensions[get_column_letter(ci)].width=12

    REF="'Consolidado Reposición'!$B$1"
    rep_cols_formula = [3+bi*3 for bi in range(N)] + [mapa2_start]

    for pi,p in enumerate(products):
        row=pi+4; rf=_row_fill(pi)
        pvt_tot=sum(p[f'{b}__pvtas'] for b in branches)
        sal_tot=sum(p[f'{b}__saldo'] for b in branches)
        avg=_avg_pvtas(p,branches)
        est=_estado_prod(pvt_tot,sal_tot,dias)

        def dc(col,val,fill,bold=False,left=False,color="000000"):
            c=ws.cell(row,col,val); c.font=_font(bold=bold,color=color,size=11)
            c.fill=_fill(fill); c.alignment=_left() if left else _center(); c.border=_border()

        dc(1,p['item'],rf,bold=True); dc(2,p['desc'],rf,left=True)
        for bi,b in enumerate(branches):
            sc2=3+bi*3; pvt=p[f'{b}__pvtas']; sal=p[f'{b}__saldo']
            rv=f"=MAX(0,ROUNDUP({pvt}*{REF}/30,0)-{int(sal)})" if pvt>0 else "-"
            dc(sc2,rv,rf); dc(sc2+1,int(sal) if sal else "-",rf); dc(sc2+2,int(pvt) if pvt>0 else "-",rf)

        mr=ws.cell(row,mapa2_start,f"=MAX(0,ROUNDUP({avg}*{REF}/30,0)-0)" if avg>0 else "-")
        mr.font=_font(bold=True,color="7F4F00",size=11); mr.fill=_fill(C_MAPA2_HD); mr.alignment=_center(); mr.border=_border()
        dc(mapa2_start+1,"-",C_MAPA2_HD,color="7F4F00"); dc(mapa2_start+2,round(avg,1) if avg>0 else "-",C_MAPA2_HD,color="7F4F00")

        rep_str=",".join(f"{get_column_letter(c)}{row}" for c in rep_cols_formula)
        tc=ws.cell(row,total_col,f"=SUM({rep_str})")
        tc.font=_font(bold=True,size=11); tc.fill=_fill(C_HDR_LIGHT); tc.alignment=_center(); tc.border=_border()
        ec=ws.cell(row,estado_col,est); ec.font=_font(bold=True,size=11)
        ec.fill=_fill(_estado_color(est)); ec.alignment=_center(); ec.border=_border()

    ws.freeze_panes="C4"

# ── HOJA 4: Detalle por Sucursal ─────────────────────────────────────────────
def sheet_detalle_sucursal(wb, products, branches, dias):
    ws = wb.create_sheet("Detalle por Sucursal")
    headers=["Código","Descripción","Sucursal","P.Vtas/Mes","Saldo","Demanda (días)","Reponer","Cobertura días","Estado"]
    for ci,h in enumerate(headers,1):
        _hdr(ws,f"{get_column_letter(ci)}1",h)
    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=48
    ws.column_dimensions["C"].width=18
    for ci in range(4,len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width=14

    REF="'Consolidado Reposición'!$B$1"
    row=2
    for p in products:
        for b in branches + ["11 Mapa 2"]:
            is_mapa2 = (b=="11 Mapa 2")
            pvt = _avg_pvtas(p,branches) if is_mapa2 else p[f'{b}__pvtas']
            sal = 0 if is_mapa2 else p[f'{b}__saldo']
            est = _estado_suc(pvt,sal,dias)
            ec  = _estado_color(est)
            rf  = C_MAPA2_HD if is_mapa2 else (_row_fill(row%2))

            if pvt>0:
                dem=f"=ROUNDUP({pvt}*{REF}/30,0)"
                rep=f"=MAX(0,ROUNDUP({pvt}*{REF}/30,0)-{int(sal)})"
                cob=f'=IF({pvt}>0,{int(sal)}/({pvt}/30),"N/A")'
            else:
                dem=0; rep=0; cob="N/A"

            vals=[p['item'],p['desc'],b,
                  int(pvt) if pvt>0 else "-",
                  int(sal) if (sal or not is_mapa2) else "-",
                  dem,rep,cob,est]
            for ci,v in enumerate(vals,1):
                c=ws.cell(row,ci,v); c.font=_font(bold=(ci==len(vals)),size=11)
                c.border=_border(); c.alignment=_left() if ci==2 else _center()
                c.fill=_fill(ec if ci==len(vals) else rf)
            row+=1

    ws.freeze_panes="A2"

# ── HOJA 5: Saldos de Bodega ──────────────────────────────────────────────────
def sheet_saldos(wb, products, branches):
    ws = wb.create_sheet("Saldos de Bodega")
    N = len(branches)
    ws["A1"].value="SALDOS DE BODEGA"; ws["A1"].font=_font(bold=True,size=13)
    _hdr(ws,"A2","Código"); _hdr(ws,"B2","Descripción")
    for bi,b in enumerate(branches):
        _hdr(ws,f"{get_column_letter(3+bi)}2",b)
    mapa2_col=3+N; total_col=mapa2_col+1
    ws[f"{get_column_letter(mapa2_col)}2"].value="11 Mapa 2"
    ws[f"{get_column_letter(mapa2_col)}2"].font=_font(bold=True,color="7F4F00",size=10)
    ws[f"{get_column_letter(mapa2_col)}2"].fill=_fill(C_MAPA2_HD)
    ws[f"{get_column_letter(mapa2_col)}2"].alignment=_center()
    ws[f"{get_column_letter(mapa2_col)}2"].border=_border()
    _hdr(ws,f"{get_column_letter(total_col)}2","Total Saldo")
    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=48
    for ci in range(3,total_col+1):
        ws.column_dimensions[get_column_letter(ci)].width=13

    for pi,p in enumerate(products):
        row=pi+3; rf=_row_fill(pi)
        def dc(col,val,fill,bold=False,left=False):
            c=ws.cell(row,col,val); c.font=_font(bold=bold,size=11)
            c.fill=_fill(fill); c.alignment=_left() if left else _center(); c.border=_border()
        dc(1,p['item'],rf,bold=True); dc(2,p['desc'],rf,left=True)
        total=0
        for bi,b in enumerate(branches):
            sal=p[f'{b}__saldo']; total+=sal; dc(3+bi,int(sal) if sal else "-",rf)
        dc(mapa2_col,"-",C_MAPA2_HD)
        tc=ws.cell(row,total_col,int(total) if total else "-")
        tc.font=_font(bold=True,size=11); tc.fill=_fill(C_HDR_LIGHT); tc.alignment=_center(); tc.border=_border()
    ws.freeze_panes="A3"

# ── HOJA 6: Sugerido Inventario Mapa 2 ───────────────────────────────────────
def sheet_mapa2(wb, products, branches, dias):
    ws = wb.create_sheet("Sugerido Inventario Mapa 2")
    ws["A1"].value="INVENTARIO SUGERIDO - NUEVA BODEGA MAPA 2"
    ws["A1"].font=_font(bold=True,size=13,color="7F4F00")
    ws["E1"].value="Días:"; ws["E1"].font=_font(bold=True,size=11)
    ws["F1"].value="='Consolidado Reposición'!B1"
    ws["F1"].font=_font(bold=True,color="7F4F00",size=11)
    ws["F1"].fill=_fill(C_YELLOW); ws["F1"].alignment=_center(); ws["F1"].border=_border()
    ws["A2"].value="Basado en ventas promedio de locales existentes. Ajusta los días según tu criterio."
    ws["A2"].font=_font(size=10,color="7F7F7F")

    for coord,val,bg,fc in [
        ("A3","Código",C_HDR_DARK,"FFFFFF"),
        ("B3","Descripción",C_HDR_DARK,"FFFFFF"),
        ("C3","Vtas Prom/Mes\n(Avg otros locales)",C_HDR_DARK,"FFFFFF"),
        ("D3","Locales\ncon ventas",C_HDR_DARK,"FFFFFF"),
        ("E3","Saldo\nActual Mapa2",C_HDR_DARK,"FFFFFF"),
        ("F3","Sugerido\n30 días",C_MAPA2_HD,"7F4F00"),
        ("G3","Sugerido\n45 días",C_MAPA2_HD,"7F4F00"),
        ("H3","Sugerido\n60 días",C_MAPA2_HD,"7F4F00"),
        ("I3","Sugerido\n(días F1)",C_MAPA2_DIN,"7F4F00"),
        ("J3","Estado rotación",C_MAPA2_HD,"7F4F00"),
    ]:
        ws[coord].value=val; ws[coord].font=_font(bold=True,color=fc,size=10)
        ws[coord].fill=_fill(bg); ws[coord].alignment=_center(); ws[coord].border=_border()
    ws.row_dimensions[3].height=42

    ws.column_dimensions["A"].width=11; ws.column_dimensions["B"].width=48
    ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=12
    ws.column_dimensions["E"].width=14
    for col in ["F","G","H","I","J"]: ws.column_dimensions[col].width=14

    row=4
    for pi,p in enumerate(products):
        avg=_avg_pvtas(p,branches)
        if avg==0: continue
        locales=_locales_venta(p,branches); rf=_row_fill(row%2)

        def dc(col,val,fill,bold=False,left=False,color="000000"):
            c=ws.cell(row,col,val); c.font=_font(bold=bold,color=color,size=11)
            c.fill=_fill(fill); c.alignment=_left() if left else _center(); c.border=_border()

        dc(1,p['item'],rf,bold=True); dc(2,p['desc'],rf,left=True)
        dc(3,avg,rf); dc(4,locales,rf); dc(5,0,rf)
        for ci,d in enumerate([30,45,60],6):
            sug=max(0,math.ceil(avg*d/30))
            c=ws.cell(row,ci,sug); c.font=_font(bold=True,color="7F4F00",size=11)
            c.fill=_fill(C_MAPA2_HD); c.alignment=_center(); c.border=_border()
        dyn=ws.cell(row,9,f"=MAX(0,ROUNDUP({avg}*'Consolidado Reposición'!$B$1/30,0)-0)")
        dyn.font=_font(bold=True,color="7F4F00",size=11); dyn.fill=_fill(C_MAPA2_DIN)
        dyn.alignment=_center(); dyn.border=_border()
        est="OK" if avg>=3 else "BAJA ROTACIÓN"
        ec=ws.cell(row,10,est); ec.font=_font(bold=True,color="7F4F00",size=11)
        ec.fill=_fill(C_MAPA2_HD); ec.alignment=_center(); ec.border=_border()
        row+=1

    ws.cell(row+1,1,f"Total productos con ventas: {row-4} de {len(products)}")
    ws.cell(row+1,1).font=_font(bold=True,size=11)
    ws.freeze_panes="A4"

# ── GENERADOR PRINCIPAL ───────────────────────────────────────────────────────
def generar_reposicion(filepath, dias=30, output_path="Reposicion.xlsx"):
    products, branches = parse_xls(filepath)
    wb = Workbook(); wb.remove(wb.active)
    sheet_consolidado(wb, products, branches, dias)
    sheet_rep_saldo(wb, products, branches, dias)
    sheet_detalle_expandido(wb, products, branches, dias)
    sheet_detalle_sucursal(wb, products, branches, dias)
    sheet_saldos(wb, products, branches)
    sheet_mapa2(wb, products, branches, dias)
    wb.save(output_path)
    return output_path, len(products), branches
